from openpyxl import load_workbook
from collections import Counter
import json, re, unicodedata

path = r"D:\Downloads\T9 (2).xlsx"
wb = load_workbook(path, data_only=False)

def color(cell):
    fill = cell.fill
    fg = fill.fgColor
    if fill.fill_type is None:
        return None
    if fg.type == "rgb":
        return (fg.rgb or "")[-6:].upper()
    if fg.type == "indexed":
        return f"INDEXED:{fg.indexed}"
    if fg.type == "theme":
        return f"THEME:{fg.theme}:{fg.tint}"
    return None

def norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    return "".join(c for c in text if unicodedata.category(c) != "Mn").replace("Đ", "D").replace("đ", "d").upper().strip()

result = {"sheets": wb.sheetnames, "matches": [], "colorCounts": {}, "sample": []}
for ws in wb.worksheets:
    counts = Counter()
    for row in ws.iter_rows():
        for cell in row:
            c = color(cell)
            if c:
                counts[c] += 1
            if "SU HAO" in norm(cell.value):
                result["matches"].append({"sheet": ws.title, "cell": cell.coordinate, "value": cell.value, "color": c})
            if norm(cell.value) in {"TIEN","HAN","QUYNH","DUONG","THUY","NGOC","NGAN","MAI","DIEM"}:
                result["matches"].append({"sheet": ws.title, "cell": cell.coordinate, "value": cell.value, "color": c})
    result["colorCounts"][ws.title] = counts.most_common(30)
    if ws.title == "Trang tính2":
        for row in range(1, 8):
            result["sample"].append([{"cell": ws.cell(row, col).coordinate, "value": ws.cell(row, col).value, "color": color(ws.cell(row, col))} for col in range(1, 15)])
        result["legend"] = [[{"cell": ws.cell(row, col).coordinate, "value": ws.cell(row, col).value, "color": color(ws.cell(row, col))} for col in range(3, 6)] for row in range(11, 20)]
print(json.dumps(result, ensure_ascii=False, indent=2))

# Compare every generated web template against the source cell named in its note.
source = open(r"D:\2DUnityGame\quocvinhvku.github.io\happychild\js\t9-templates.js", encoding="utf-8").read()
teachers = json.loads(re.search(r"export const T9_TEACHER_SOURCE=(\[.*?\]);\s*export const T9_TEMPLATE_SOURCE", source, re.S).group(1))
templates = json.loads(re.search(r"export const T9_TEMPLATE_SOURCE=(\[.*\]);\s*$", source, re.S).group(1))
teacher_colors = {t["id"]: t["color"].lstrip("#").upper() for t in teachers}
aliases = {
    "EA9999": "F48E93", # older pink tint used for Cô Quỳnh
    "548235": "38761D", # older dark-green tint used for Cô Hân
    "FF0000": "CC0000", # red tint used for Cô Mai
}
mismatches = []
for item in templates:
    match = re.search(r"·\s*([^·]+?)\s*·\s*([A-Z]+\d+)\s*$", item.get("note", ""))
    if not match:
        continue
    sheet_name, coord = match.group(1).strip(), match.group(2)
    if sheet_name not in wb.sheetnames:
        continue
    actual = color(wb[sheet_name][coord])
    expected = teacher_colors.get(item.get("teacherId"))
    normalized_actual = aliases.get(actual, actual)
    if actual and actual not in ("FFFFFF", "000000", "E7E6E6", "F2F2F2") and normalized_actual != expected:
        mismatches.append({"id": item.get("id"), "student": item.get("studentName"), "cell": coord, "time": wb[sheet_name][coord].value, "actualColor": actual, "webTeacher": item.get("teacherId"), "expectedColor": expected})
print(json.dumps({"coloredTemplates": len(templates), "mismatches": mismatches}, ensure_ascii=False, indent=2))
